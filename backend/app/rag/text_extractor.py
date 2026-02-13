"""
Extraction de texte depuis différents formats de fichiers.
Supporte : PDF, Word (.docx), Excel (.xlsx), images (OCR via pytesseract).
"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)


async def extract_text_from_file(file_path: str, mime_type: str | None = None) -> str:
    """
    Extrait le texte brut d'un fichier selon son type MIME ou son extension.
    Retourne le texte extrait ou une chaîne vide en cas d'échec.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {file_path}")

    # Déterminer le type à partir du mime ou de l'extension
    mime = (mime_type or "").lower()
    suffix = path.suffix.lower()

    if mime == "application/pdf" or suffix == ".pdf":
        return _extract_pdf(path)
    elif mime in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ) or suffix in (".docx", ".doc"):
        return _extract_docx(path)
    elif mime in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ) or suffix in (".xlsx", ".xls"):
        return _extract_xlsx(path)
    elif mime.startswith("image/") or suffix in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
        return _extract_image_ocr(path)
    else:
        # Tenter de lire comme texte brut
        try:
            return path.read_text(encoding="utf-8")
        except Exception:
            raise ValueError(f"Type de fichier non supporté : {mime or suffix}")


def _extract_pdf(path: Path) -> str:
    """Extrait le texte d'un PDF via pdfplumber."""
    import pdfplumber

    pages_text = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                pages_text.append(f"[Page {i + 1}]\n{text}")
    return "\n\n".join(pages_text)


def _extract_docx(path: Path) -> str:
    """Extrait le texte d'un fichier Word (.docx)."""
    from docx import Document

    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_xlsx(path: Path) -> str:
    """Extrait le texte d'un fichier Excel (.xlsx) — toutes les feuilles."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True)
    sheets_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            sheets_text.append(f"[Feuille: {sheet_name}]\n" + "\n".join(rows))
    return "\n\n".join(sheets_text)


def _extract_image_ocr(path: Path) -> str:
    """Extrait le texte d'une image via OCR (pytesseract)."""
    try:
        import pytesseract
        from PIL import Image

        image = Image.open(path)
        text = pytesseract.image_to_string(image, lang="fra+eng")
        return text.strip()
    except ImportError:
        logger.warning("pytesseract ou Pillow non installé — OCR indisponible")
        return ""
    except Exception as e:
        logger.warning("Erreur OCR sur %s : %s", path, e)
        return ""
